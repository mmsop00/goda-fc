import openpyxl
import json

wb = openpyxl.load_workbook(r"c:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx")
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} === (rows: {ws.max_row}, cols: {ws.max_column})")
    for row in ws.iter_rows(max_row=min(ws.max_row, 25), values_only=True):
        print(" | ".join([str(c or "") for c in row]))
