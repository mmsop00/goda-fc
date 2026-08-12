import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)

# Check "Tỷ số" sheet
if 'Tỷ số' in wb.sheetnames:
    ws = wb['Tỷ số']
    print('\n=== Tỷ số sheet ===')
    # Print headers
    headers = [str(c.value) for c in ws[1]]
    for i, h in enumerate(headers):
        print(f"  Col {i}: {h}")
    
    print('\n--- Data rows ---')
    for r_idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not r[0] and not r[1]:
            continue
        print(f"Row {r_idx}: {[str(c)[:40] if c else 'None' for c in r[:15]]}")
