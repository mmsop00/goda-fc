import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)
ws = wb['Thông tin cầu thủ']

# Print headers to find the right columns
headers = [str(c.value) for c in ws[1]]
for i, h in enumerate(headers):
    print(f"  Col {i}: {h}")

print("\n--- Data ---")
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]:
        continue
    name = str(r[1]).strip()
    number = r[7] if r[7] else ""
    shirt_name = r[8] if r[8] else ""
    size = r[9] if r[9] else ""
    set2 = r[10] if r[10] else ""
    position = r[5] if r[5] else ""
    print(f"{name:<25s} | #{str(number):<4s} | In áo: {str(shirt_name):<20s} | Size: {str(size):<6s} | Bộ 2: {str(set2):<6s} | VT: {str(position)}")
