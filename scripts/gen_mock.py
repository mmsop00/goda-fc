import openpyxl, re, json
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)

# ─── GODA player names ───
ws_players = wb['Thông tin cầu thủ']
goda_names = set()
for row in ws_players.iter_rows(min_row=2, values_only=True):
    if row[1]:
        goda_names.add(str(row[1]).strip().lower())
    if row[8]:
        goda_names.add(str(row[8]).strip().lower())

def is_goda(name):
    return name.lower() in goda_names

def parse_goals(goals_str):
    if not goals_str or goals_str in ('None', '', ' '):
        return []
    result = []
    parts = re.split(r'\s*-\s*', goals_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        m = re.match(r"(\d+\+?\d*)\s*'?\s*:?\s*(.+?)(?:,\s*(.+))?$", part)
        if m:
            minute = int(re.sub(r'\+.*', '', m.group(1)))
            player = m.group(2).strip()
            assist = (m.group(3) or '').strip()
            if player.lower() == 'unknown' or player == '':
                continue
            if assist.lower() == 'unknown':
                assist = ''
            result.append({'player': player, 'minute': minute, 'assist': assist})
    return result

# ─── Parse Matches ───
ws = wb['Tỷ số']
matches = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    date_raw = str(row[1]).strip() if row[1] else ''
    time, date = '', ''
    if ' - ' in date_raw:
        p = date_raw.split(' - ')
        time, date = p[0].strip(), p[1].strip()
    else:
        date = date_raw
    
    venue = str(row[3]).strip() if row[3] else ''
    home_team = str(row[4]).strip() if row[4] else ''
    away_team = str(row[5]).strip() if row[5] else ''
    score_raw = str(row[6]).strip() if row[6] else ''
    home_goals_raw = str(row[7]).strip() if row[7] else ''
    away_goals_raw = str(row[8]).strip() if row[8] else ''
    maps_url = str(row[9]).strip() if row[9] else ''
    
    hs, aws = 0, 0
    if ' - ' in score_raw:
        sp = score_raw.split(' - ')
        try:
            hs, aws = int(sp[0].strip()), int(sp[1].strip())
        except: pass
    
    is_home = 'GODA' in home_team
    opponent = away_team if is_home else home_team
    gs = hs if is_home else aws
    ops = aws if is_home else hs
    
    hg = parse_goals(home_goals_raw)
    ag = parse_goals(away_goals_raw)
    
    goda_goals, opp_goals = [], []
    
    # Assign sides, detect GODA players even in wrong column
    for g in hg:
        if is_home or is_goda(g['player']):
            g['side'] = 'GODA'; goda_goals.append(g)
        else:
            g['side'] = 'opponent'; opp_goals.append(g)
    for g in ag:
        if (not is_home) or is_goda(g['player']):
            g['side'] = 'GODA'; goda_goals.append(g)
        else:
            g['side'] = 'opponent'; opp_goals.append(g)
    
    all_goals = goda_goals + opp_goals
    all_goals.sort(key=lambda g: g['minute'])
    
    # MVP from GODA scorers
    mvp = goda_goals[-1]['player'] if goda_goals else ''
    
    matches.append({
        'date': date, 'time': time, 'venue': venue,
        'isHome': is_home, 'opponent': opponent,
        'godaScore': gs, 'opponentScore': ops,
        'goals': all_goals, 'mvp': mvp,
        'googleMapsUrl': maps_url,
    })

# ─── Output MOCK_MATCH_RESULTS ───
print("// ═══ MOCK_MATCH_RESULTS ═══")
print("export const MOCK_MATCH_RESULTS: MatchResult[] = [")
for i, m in enumerate(matches):
    gts = json.dumps(m['goals'], ensure_ascii=False)
    mvp_line = f',\n    mvp: "{m["mvp"]}"' if m['mvp'] else ''
    print(f'''  {{
    id: "mr-{i+1:03d}",
    season: "2026",
    date: "{m['date']}",
    time: "{m['time']}",
    venue: "{m['venue']}, Hà Nội",
    type: "Giao hữu",
    isHome: {'true' if m['isHome'] else 'false'},
    opponent: "{m['opponent']}",
    opponentScore: {m['opponentScore']},
    godaScore: {m['godaScore']},
    godaLineup: BASE_LINEUP,
    opponentLineup: OPP_LINEUP,
    goals: {gts},
    cards: [],
    imageUrl: "https://placehold.co/800x400/0B1E3A/F7C600?text=GODA+{m['godaScore']}-{m['opponentScore']}",
    googleMapsUrl: "{m['googleMapsUrl']}",{mvp_line}
  }},''')
print("];")

# ─── Output MOCK_MEMBERS ───
print("\n// ═══ MOCK_MEMBERS ═══")
members = []
for row in ws_players.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    name = str(row[1]).strip()
    nickname = str(row[2]).strip() if row[2] else ''
    br = row[3]
    if br:
        if isinstance(br, datetime):
            bday = br.strftime('%d/%m')
        else:
            try:
                bday = datetime.strptime(str(br)[:10], '%Y-%m-%d').strftime('%d/%m')
            except:
                bday = str(br)[:5]
    else:
        bday = ''
    num = int(row[7]) if row[7] else 0
    jersey = str(row[8]).strip() if row[8] else ''
    size = str(row[9]).strip() if row[9] else ''
    role = str(row[10]).strip() if row[10] else ''
    pos = str(row[12]).strip() if row[12] else ''
    jr = row[17]
    if jr:
        if isinstance(jr, datetime): jy = jr.year
        else:
            try: jy = int(str(jr)[:4])
            except: jy = 0
    else:
        jy = 0
    status = str(row[18]).strip() if row[18] else 'Đang thi đấu'
    av = str(row[20]).strip() if row[20] else ''
    if 'drive.google.com/file/d/' in av:
        fid = av.split('/d/')[1].split('/')[0]
        av = f'https://drive.google.com/uc?export=view&id={fid}'
    elif 'drive.google.com' in av:
        av = ''
    members.append({
        'name': name, 'nickname': nickname, 'position': pos,
        'number': num, 'avatarUrl': av, 'birthday': bday,
        'joinYear': jy, 'status': role or status,
    })

# Sort: Đội trưởng -> Đội phó -> by joinYear desc
def sk(m):
    if m['status'] == 'Đội trưởng': return (0, 0)
    if m['status'] == 'Đội phó': return (1, 0)
    return (2, -(m['joinYear'] or 0))

members.sort(key=sk)

print("export const MOCK_MEMBERS: MemberPublic[] = [")
for i, m in enumerate(members):
    comment = ''
    if m['status'] == 'Đội trưởng': comment = ' // ── Đội trưởng ──'
    elif m['status'] == 'Đội phó': comment = ' // ── Đội phó ──'
    print(f'  {{ id: "m-{i+1:03d}", name: "{m["name"]}", nickname: "{m["nickname"]}", position: "{m["position"]}", number: {m["number"]}, avatarUrl: "{m["avatarUrl"]}", matches: 0, goals: 0, assists: 0, mvp: 0, birthday: "{m["birthday"]}", joinYear: {m["joinYear"]}, status: "{m["status"]}" }},{comment}')
print("];")
