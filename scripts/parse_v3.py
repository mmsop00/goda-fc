import openpyxl, re
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)

# GODA player names for detection
ws_p = wb['Thông tin cầu thủ']
goda_names = set()
for row in ws_p.iter_rows(min_row=2, values_only=True):
    if row[1]: goda_names.add(str(row[1]).strip().lower())
    if row[8]: goda_names.add(str(row[8]).strip().lower())

def is_goda(name):
    return name.lower() in goda_names

def parse_goals(raw):
    """Parse goal string into list of {player, minute, assist, isUnknown}"""
    if not raw or str(raw).strip() in ('', 'None', ' '):
        return []
    s = str(raw).strip()
    result = []
    # Split by " - " (separates goals)
    parts = re.split(r'\s*-\s*', s)
    for part in parts:
        part = part.strip()
        if not part: continue
        # "85': Player, Asst" or "30' Player, Asst" or "20': Unknown, Unknown"
        m = re.match(r"(\d+\+?\d*)\s*'?\s*:?\s*(.+?)(?:,\s*(.+))?$", part)
        if m:
            minute_str = m.group(1).replace('+', '')
            minute = int(minute_str)
            player = m.group(2).strip()
            assist = (m.group(3) or '').strip()
            is_unknown = player.lower() == 'unknown'
            if assist.lower() == 'unknown': assist = ''
            result.append({'player': player, 'minute': minute, 'assist': assist, 'isUnknown': is_unknown})
    return result

# Parse matches
ws = wb['Tỷ số']
matches = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    
    date_raw = str(row[1]).strip() if row[1] else ''
    time, date = '', ''
    if ' - ' in date_raw:
        p = date_raw.split(' - ')
        time, date = p[0].strip(), p[1].strip()
    else:
        date = date_raw
    
    venue = str(row[3]).strip() if row[3] else ''
    home_team = str(row[4]).strip() if row[4] else ''
    away_team = str(row[5]).strip() if row[5] else ''
    score_raw = str(row[6]).strip() if row[6] else ''
    home_goals_raw = str(row[7]).strip() if row[7] else ''
    away_goals_raw = str(row[8]).strip() if row[8] else ''
    maps_url = str(row[9]).strip() if row[9] else ''
    
    # Parse score
    hs, aws = 0, 0
    if ' - ' in score_raw:
        sp = score_raw.split(' - ')
        try: hs, aws = int(sp[0].strip()), int(sp[1].strip())
        except: pass
    
    is_home = 'GODA' in home_team
    opponent = away_team if is_home else home_team
    goda_score = hs if is_home else aws
    opp_score = aws if is_home else hs
    
    # Parse goals from both columns
    col_a = parse_goals(home_goals_raw)  # "Người ghi bàn đội nhà"
    col_b = parse_goals(away_goals_raw)  # "Người ghi bàn đội khách"
    
    # Detect which column has GODA players
    goda_in_a = any(is_goda(g['player']) for g in col_a)
    goda_in_b = any(is_goda(g['player']) for g in col_b)
    
    # Determine correct assignment
    if is_home:
        expected_goda_col = 'a'
        expected_opp_col = 'b'
    else:
        expected_goda_col = 'b'
        expected_opp_col = 'a'
    
    # Check if columns are swapped by comparing goal counts to scores
    count_a = len(col_a)
    count_b = len(col_b)
    expected_goda_goals = goda_score
    expected_opp_goals = opp_score
    
    a_is_goda = (expected_goda_col == 'a')
    b_is_goda = (expected_goda_col == 'b')
    
    # If GODA players detected in wrong column, swap
    if a_is_goda and goda_in_b and not goda_in_a:
        a_is_goda, b_is_goda = False, True
    elif b_is_goda and goda_in_a and not goda_in_b:
        a_is_goda, b_is_goda = True, False
    
    # If all are Unknown, check goal counts vs score
    all_unknown = all(g['isUnknown'] for g in col_a + col_b)
    if all_unknown and expected_goda_goals > 0:
        if count_a == expected_goda_goals and a_is_goda == False:
            a_is_goda, b_is_goda = True, False
        elif count_b == expected_goda_goals and b_is_goda == False:
            a_is_goda, b_is_goda = False, True
    
    # Build final goals list
    goda_goals = []
    opp_goals = []
    
    for g in col_a:
        g = dict(g)
        if a_is_goda:
            g['side'] = 'GODA'
            if g['isUnknown']: g['player'] = 'Cầu thủ GODA FC'
            goda_goals.append(g)
        else:
            g['side'] = 'opponent'
            if g['isUnknown']: g['player'] = f'Cầu thủ {opponent}'
            opp_goals.append(g)
    
    for g in col_b:
        g = dict(g)
        if b_is_goda:
            g['side'] = 'GODA'
            if g['isUnknown']: g['player'] = 'Cầu thủ GODA FC'
            goda_goals.append(g)
        else:
            g['side'] = 'opponent'
            if g['isUnknown']: g['player'] = f'Cầu thủ {opponent}'
            opp_goals.append(g)
    
    all_goals = goda_goals + opp_goals
    all_goals.sort(key=lambda g: g['minute'])
    
    mvp = goda_goals[-1]['player'] if goda_goals and not goda_goals[-1]['isUnknown'] else ''
    
    matches.append({
        'date': date, 'time': time, 'venue': venue,
        'isHome': is_home, 'opponent': opponent,
        'godaScore': goda_score, 'opponentScore': opp_score,
        'goals': all_goals, 'mvp': mvp,
        'googleMapsUrl': maps_url,
    })
    
    print(f"Match: {date} {time} | {'GODA' if is_home else opponent} {goda_score}-{opp_score} {'GODA' if not is_home else opponent}")
    print(f"  col_a (home goals): {col_a} -> goda={a_is_goda}")
    print(f"  col_b (away goals): {col_b} -> goda={b_is_goda}")
    print(f"  GODA goals: {goda_goals}")
    print(f"  OPP goals: {opp_goals}")
    print(f"  MVP: {mvp}")
    print()

# Generate TypeScript
print("\n\n// ═══ MOCK_MATCH_RESULTS ═══")
print("export const MOCK_MATCH_RESULTS: MatchResult[] = [")
for i, m in enumerate(matches):
    goals_ts = "[\n"
    for g in m['goals']:
        a = f', assist: "{g["assist"]}"' if g['assist'] else ''
        goals_ts += f'      {{ player: "{g["player"]}", minute: {g["minute"]}{a}, side: "{g["side"]}" }},\n'
    goals_ts += "    ]"
    mvp_ts = f',\n    mvp: "{m["mvp"]}"' if m['mvp'] else ''
    print(f'''  {{
    id: "mr-{i+1:03d}",
    season: "2026",
    date: "{m['date']}",
    time: "{m['time']}",
    venue: "Sân {m['venue']}, Hà Nội",
    type: "Giao hữu",
    isHome: {'true' if m['isHome'] else 'false'},
    opponent: "{m['opponent']}",
    opponentScore: {m['opponentScore']},
    godaScore: {m['godaScore']},
    godaLineup: BASE_LINEUP,
    opponentLineup: OPP_LINEUP,
    goals: {goals_ts},
    cards: [],
    imageUrl: "https://placehold.co/800x400/0B1E3A/F7C600?text=GODA+{m['godaScore']}-{m['opponentScore']}",{mvp_ts}
    googleMapsUrl: "{m['googleMapsUrl']}",
  }},''')
print("];")
