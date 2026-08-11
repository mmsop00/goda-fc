import openpyxl
import json
import re
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)

# ─── Parse Players ───
ws_players = wb['Thông tin cầu thủ']
players = []
for row in ws_players.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    name = str(row[1]).strip() if row[1] else ''
    nickname = str(row[2]).strip() if row[2] else ''
    birthday_raw = row[3]
    if birthday_raw:
        if isinstance(birthday_raw, datetime):
            birthday = birthday_raw.strftime('%d/%m')
        else:
            bstr = str(birthday_raw).strip()
            try:
                dt = datetime.strptime(bstr[:10], '%Y-%m-%d')
                birthday = dt.strftime('%d/%m')
            except:
                birthday = bstr[:5] if len(bstr) >= 5 else bstr
    else:
        birthday = ''
    
    number = int(row[7]) if row[7] else 0
    jersey_name = str(row[8]).strip() if row[8] else ''
    size = str(row[9]).strip() if row[9] else ''
    role = str(row[10]).strip() if row[10] else ''
    position = str(row[12]).strip() if row[12] else ''
    join_raw = row[17]
    if join_raw:
        if isinstance(join_raw, datetime):
            join_year = join_raw.year
        else:
            try:
                join_year = int(str(join_raw)[:4])
            except:
                join_year = 0
    else:
        join_year = 0
    
    status = str(row[18]).strip() if row[18] else ''
    avatar = str(row[20]).strip() if row[20] else ''
    
    players.append({
        'name': name,
        'nickname': nickname,
        'birthday': birthday,
        'number': number,
        'jerseyName': jersey_name,
        'size': size,
        'role': role,
        'position': position,
        'joinYear': join_year,
        'status': status,
        'avatarUrl': avatar,
    })

print(f"=== PLAYERS ({len(players)}) ===")
for i, p in enumerate(players):
    print(f"[{i}] {p['name']} | {p['nickname']} | #{p['number']} | {p['position']} | {p['status']} | join={p['joinYear']} | bday={p['birthday']} | jersey={p['jerseyName']} | size={p['size']} | role={p['role']} | avatar={p['avatarUrl'][:60] if p['avatarUrl'] else 'NONE'}")

print("\n=== MATCHES ===")
# ─── Parse Matches ───
ws_matches = wb['Tỷ số']
matches = []
for row in ws_matches.iter_rows(min_row=2, values_only=True):
    stt = row[0]
    if not stt:
        continue
    
    date_raw = str(row[1]).strip() if row[1] else ''
    # Parse "16h00 - 18/07/2026" format
    time = ''
    date = ''
    if ' - ' in date_raw:
        parts = date_raw.split(' - ')
        time = parts[0].strip()
        date = parts[1].strip()
    else:
        date = date_raw
    
    status = str(row[2]).strip() if row[2] else ''
    venue = str(row[3]).strip() if row[3] else ''
    home_team = str(row[4]).strip() if row[4] else ''
    away_team = str(row[5]).strip() if row[5] else ''
    score_raw = str(row[6]).strip() if row[6] else ''
    home_goals_str = str(row[7]).strip() if row[7] else ''
    away_goals_str = str(row[8]).strip() if row[8] else ''
    maps_url = str(row[9]).strip() if row[9] else ''
    color = str(row[10]).strip() if row[10] else ''
    
    # Parse score
    home_score = 0
    away_score = 0
    if ' - ' in score_raw:
        sp = score_raw.split(' - ')
        try:
            home_score = int(sp[0].strip())
            away_score = int(sp[1].strip())
        except:
            pass
    
    is_home = 'GODA' in home_team
    opponent = away_team if is_home else home_team
    goda_score = home_score if is_home else away_score
    opp_score = away_score if is_home else home_score
    
    # Parse goal scorers
    def parse_goals(goals_str, side):
        if not goals_str or goals_str == 'None' or goals_str == '':
            return []
        result = []
        # Split by " - " which separates goals
        parts = goals_str.split(' - ')
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # Pattern: "85': Player Name, Assistant" or "30' Player Name, Assistant" or "85': Player Name"
            match = re.match(r"(\d+\+?\d*)\s*'?:\s*(.+?)(?:,\s*(.+))?$", part)
            if not match:
                match = re.match(r"(\d+\+?\d*)\s*'\s*(.+?)(?:,\s*(.+))?$", part)
            if match:
                minute_str = match.group(1).replace('+', '')
                try:
                    minute = int(minute_str)
                except:
                    minute = 0
                player = match.group(2).strip()
                assist = match.group(3).strip() if match.group(3) else ''
                
                # Filter out "Unknown"
                if player.lower() == 'unknown' or player == '':
                    continue
                if assist.lower() == 'unknown':
                    assist = ''
                
                result.append({
                    'player': player,
                    'minute': minute,
                    'assist': assist,
                    'side': side,
                })
        return result
    
    # Goda goals = whoever GODA is
    goda_goals_raw = home_goals_str if is_home else away_goals_str
    opp_goals_raw = away_goals_str if is_home else home_goals_str
    
    goda_goals = parse_goals(goda_goals_raw, 'GODA')
    opp_goals = parse_goals(opp_goals_raw, 'opponent')
    
    all_goals = goda_goals + opp_goals
    # Sort by minute
    all_goals.sort(key=lambda g: g['minute'])
    
    match_type = 'Giao hữu'
    
    matches.append({
        'date': date,
        'time': time,
        'venue': venue,
        'type': match_type,
        'isHome': is_home,
        'opponent': opponent,
        'godaScore': goda_score,
        'opponentScore': opp_score,
        'goals': all_goals,
        'googleMapsUrl': maps_url,
        'color': color,
    })
    
    print(f"STT {stt}: {date} {time} | {home_team} {home_score}-{away_score} {away_team}")
    print(f"  isHome={is_home} opponent={opponent} goda={goda_score} opp={opp_score}")
    print(f"  GODA goals: {goda_goals}")
    print(f"  OPP goals: {opp_goals}")
    print(f"  venue={venue} maps={maps_url}")
    print()

print(f"\nTotal matches: {len(matches)}")
