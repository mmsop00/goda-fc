import openpyxl
import re
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\mmsop\OneDrive\Desktop\GODA FC\GODA FC - Danh sách thành viên.xlsx', data_only=True)

# ─── Parse all players first to get GODA player names ───
ws_players = wb['Thông tin cầu thủ']
goda_player_names = set()
for row in ws_players.iter_rows(min_row=2, values_only=True):
    if row[1]:
        name = str(row[1]).strip()
        goda_player_names.add(name.lower())
        # Also add jersey names
        if row[8]:
            goda_player_names.add(str(row[8]).strip().lower())

def is_goda_player(name):
    return name.lower() in goda_player_names

def parse_goals(goals_str):
    """Parse goal string into list of {player, minute, assist}"""
    if not goals_str or goals_str in ('None', '', ' '):
        return []
    result = []
    # Split by " - " between goals
    parts = re.split(r'\s*-\s*', goals_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Pattern: "85': Player Name, Assistant" or "30' Player Name, Assistant" or "85': Player Name"
        m = re.match(r"(\d+\+?\d*)\s*'?\s*:?\s*(.+?)(?:,\s*(.+))?$", part)
        if m:
            minute = int(re.sub(r'\+.*', '', m.group(1)))
            player = m.group(2).strip()
            assist = (m.group(3) or '').strip()
            if player.lower() == 'unknown' or player == '':
                continue
            if assist.lower() == 'unknown':
                assist = ''
            result.append({'player': player, 'minute': minute, 'assist': assist})
    return result

# ─── Parse Matches ───
ws_matches = wb['Tỷ số']
matches = []
for row in ws_matches.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    
    date_raw = str(row[1]).strip() if row[1] else ''
    time = ''
    date = ''
    if ' - ' in date_raw:
        parts = date_raw.split(' - ')
        time = parts[0].strip()
        date = parts[1].strip()
    else:
        date = date_raw
    
    status = str(row[2]).strip() if row[2] else ''
    venue = str(row[3]).strip() if row[3] else ''
    home_team = str(row[4]).strip() if row[4] else ''
    away_team = str(row[5]).strip() if row[5] else ''
    score_raw = str(row[6]).strip() if row[6] else ''
    home_goals_raw = str(row[7]).strip() if row[7] else ''
    away_goals_raw = str(row[8]).strip() if row[8] else ''
    maps_url = str(row[9]).strip() if row[9] else ''
    
    # Parse score
    home_score = 0
    away_score = 0
    if ' - ' in score_raw:
        sp = score_raw.split(' - ')
        try:
            home_score = int(sp[0].strip())
            away_score = int(sp[1].strip())
        except:
            pass
    
    is_home = 'GODA' in home_team
    opponent = away_team if is_home else home_team
    goda_score = home_score if is_home else away_score
    opp_score = away_score if is_home else home_score
    
    # Parse goals from both columns
    home_goals = parse_goals(home_goals_raw)
    away_goals = parse_goals(away_goals_raw)
    
    # Determine which goals belong to GODA
    # The user may have put GODA goals in wrong column, so detect by player names
    goda_goals = []
    opp_goals = []
    
    col_a = home_goals  # "Người ghi bàn đội nhà"
    col_b = away_goals  # "Người ghi bàn đội khách"
    
    if is_home:
        # GODA is home: col_a = GODA goals, col_b = opponent goals
        for g in col_a:
            g['side'] = 'GODA'
            goda_goals.append(g)
        for g in col_b:
            # Check if player is actually a GODA player (data entry error)
            if is_goda_player(g['player']):
                g['side'] = 'GODA'
                goda_goals.append(g)
            else:
                g['side'] = 'opponent'
                opp_goals.append(g)
    else:
        # GODA is away: col_b = GODA goals, col_a = opponent goals
        for g in col_b:
            g['side'] = 'GODA'
            goda_goals.append(g)
        for g in col_a:
            # Check if player is actually a GODA player (data entry error)
            if is_goda_player(g['player']):
                g['side'] = 'GODA'
                goda_goals.append(g)
            else:
                g['side'] = 'opponent'
                opp_goals.append(g)
    
    all_goals = goda_goals + opp_goals
    all_goals.sort(key=lambda g: g['minute'])
    
    # Filter out "Unknown" or empty
    all_goals = [g for g in all_goals if g['player'].lower() != 'unknown' and g['player']]
    
    match_type = 'Giao hữu'
    
    # Determine MVP from GODA goals if any
    mvp = ''
    if goda_goals:
        # Last GODA goal scorer is often MVP candidate
        pass
    
    matches.append({
        'date': date,
        'time': time,
        'venue': venue,
        'type': match_type,
        'isHome': is_home,
        'opponent': opponent,
        'godaScore': goda_score,
        'opponentScore': opp_score,
        'goals': all_goals,
        'googleMapsUrl': maps_url,
    })
    
    print(f"  {date} {time} | {'GODA' if is_home else opponent} vs {'GODA' if not is_home else opponent} | {goda_score}-{opp_score}")
    print(f"  GODA goals: {goda_goals}")
    print(f"  OPP goals: {opp_goals}")

print(f"\nTotal matches: {len(matches)}")

# ─── Generate TypeScript for mock-data.ts ───
print("\n\n// ═══ GENERATED TYPESCRIPT ═══")
print("\n// ── MOCK_MATCH_RESULTS ──")
print("export const MOCK_MATCH_RESULTS: MatchResult[] = [")
for i, m in enumerate(matches):
    id_str = f"mr-{i+1:03d}"
    goals_ts = "[\n"
    for g in m['goals']:
        assist_ts = f', assist: "{g["assist"]}"' if g['assist'] else ''
        goals_ts += f'      {{ player: "{g["player"]}", minute: {g["minute"]}{assist_ts}, side: "{g["side"]}" }},\n'
    goals_ts += "    ]"
    
    image_url = f'https://placehold.co/800x400/0B1E3A/F7C600?text=GODA+{m["godaScore"]}-{m["opponentScore"]}'
    
    print(f'''  {{
    id: "{id_str}",
    season: "2026",
    date: "{m['date']}",
    time: "{m['time']}",
    venue: "{m['venue']}, Hà Nội",
    type: "{m['type']}",
    isHome: {'true' if m['isHome'] else 'false'},
    opponent: "{m['opponent']}",
    opponentScore: {m['opponentScore']},
    godaScore: {m['godaScore']},
    godaLineup: BASE_LINEUP,
    opponentLineup: OPP_LINEUP,
    goals: {goals_ts},
    cards: [],
    imageUrl: "{image_url}",
    googleMapsUrl: "{m['googleMapsUrl']}",
  }},''')

print("];")

# ── MOCK_MEMBERS ──
print("\n// ── MOCK_MEMBERS ──")
print("export const MOCK_MEMBERS: MemberPublic[] = [")
# Sort: Đội trưởng first, then Đội phó, then by joinYear (oldest first)
# Build players list from ws_players
players = []
for row in ws_players.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    name = str(row[1]).strip() if row[1] else ''
    nickname = str(row[2]).strip() if row[2] else ''
    birthday_raw = row[3]
    if birthday_raw:
        if isinstance(birthday_raw, datetime):
            birthday = birthday_raw.strftime('%d/%m')
        else:
            bstr = str(birthday_raw).strip()
            try:
                dt = datetime.strptime(bstr[:10], '%Y-%m-%d')
                birthday = dt.strftime('%d/%m')
            except:
                birthday = bstr[:5] if len(bstr) >= 5 else bstr
    else:
        birthday = ''
    number = int(row[7]) if row[7] else 0
    jersey_name = str(row[8]).strip() if row[8] else ''
    size = str(row[9]).strip() if row[9] else ''
    role = str(row[10]).strip() if row[10] else ''
    position = str(row[12]).strip() if row[12] else ''
    join_raw = row[17]
    if join_raw:
        if isinstance(join_raw, datetime):
            join_year = join_raw.year
        else:
            try:
                join_year = int(str(join_raw)[:4])
            except:
                join_year = 0
    else:
        join_year = 0
    status = str(row[18]).strip() if row[18] else ''
    avatar = str(row[20]).strip() if row[20] else ''
    players.append({
        'name': name, 'nickname': nickname, 'birthday': birthday,
        'number': number, 'jerseyName': jersey_name, 'size': size,
        'role': role, 'position': position, 'joinYear': join_year,
        'status': status, 'avatarUrl': avatar,
    })

def sort_key(p):
    if p['role'] == 'Đội trưởng':
        return (0, 0)
    elif p['role'] == 'Đội phó':
        return (1, 0)
    else:
        return (2, -(p['joinYear'] or 0))

players_sorted = sorted(players, key=sort_key)

for i, p in enumerate(players_sorted):
    id_str = f"m-{i+1:03d}"
    avatar = p['avatarUrl']
    # Convert Google Drive URL to direct link
    if 'drive.google.com/file/d/' in avatar:
        file_id = avatar.split('/d/')[1].split('/')[0]
        avatar = f'https://drive.google.com/uc?export=view&id={file_id}'
    elif 'drive.google.com' in avatar:
        avatar = ''
    
    role_comment = ''
    if p['role'] == 'Đội trưởng':
        role_comment = ' // ── Đội trưởng ──'
    elif p['role'] == 'Đội phó':
        role_comment = ' // ── Đội phó ──'
    
    status_display = p['status'] if p['status'] else 'Đang thi đấu'
    
    print(f'''  {{ id: "{id_str}", name: "{p['name']}", nickname: "{p['nickname']}", position: "{p['position']}", number: {p['number']}, avatarUrl: "{avatar}", matches: 0, goals: 0, assists: 0, mvp: 0, birthday: "{p['birthday']}", joinYear: {p['joinYear']}, status: "{p['role'] or status_display}" }},{role_comment}''')

print("];")
